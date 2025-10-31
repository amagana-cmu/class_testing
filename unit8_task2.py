import requests
from requests.exceptions import RequestException
from bs4 import BeautifulSoup
from bs4.element import NavigableString
import json
import csv
import xml.etree.ElementTree as ET
from typing import List, Dict, Any, Optional

from openpyxl import Workbook

def get_names_page(num: int) -> Optional[str]:
    """
    Fetches the HTML content for a given page number from behindthename.com.
    
    Args:
        num: The page number to fetch.
        
    Returns:
        The HTML content as a string, or None if an error occurs.
    """
    # Construct the URL for the specific page number
    url = f"https://www.behindthename.com/names/{num}"
    
    # Set a User-Agent to be a good web citizen
    headers = {
        'User-Agent': 'YourAppName/1.0 (YourSchoolOrPersonalProject; your-email@example.com)'
    }
    
    try:
        response = requests.get(url, headers=headers)
        # Raise an exception for bad status codes (4xx or 5xx)
        response.raise_for_status()
        return response.text
    except RequestException as e:
        print(f"Error fetching page {num}: {e}")
        return None

def extract_page_count(page_text: str) -> int:
    soup = BeautifulSoup(page_text, 'html.parser')
    
    max_page = 0
    pagination = soup.find('nav', class_='pagination')
    
    if pagination:
        # Find all links within the pagination div
        links = pagination.find_all('a')
        
        for link in links:
            # Check if the link's text is a number
            if link.text.isdigit():
                page_num = int(link.text)
                if page_num > max_page:
                    max_page = page_num

    return max_page

def extract_names_from_page(page_text: str, gender: bool = True, usage: bool = False, desc: bool = False) -> Dict[str, Dict[str, Any]]:
    """
    Parses the HTML text of a page and extracts name data based on parameters.
    
    Args:
        page_text: The HTML content of a page.
        gender: Include gender information.
        usage: Include usage (origin/language) information.
        desc: Include description information.
        
    Returns:
        A dictionary where keys are names and values are dicts of their data.
    """
    soup = BeautifulSoup(page_text, 'html.parser')
    results: Dict[str, Dict[str, Any]] = {}
    
    # Find all name entries, which are marked by a span with class "listname"
    name_spans = soup.find_all('span', class_='listname')
    
    for name_span in name_spans:
        # Get the name (text inside the <span>)
        name = name_span.text.strip()
        
        # The parent <div> contains the siblings
        parent_div = name_span.parent
        
        name_data: Dict[str, Any] = {}
        
        # Process Gender
        if gender:
            gender_span = name_span.find_next_sibling('span', class_='listgender')
            if gender_span:
                gender_text = gender_span.text.strip().lower()
                # Apply the specific logic from your prompt
                if 'masculine' in gender_text and 'feminine' in gender_text:
                    name_data['gender'] = 'm & f'
                elif 'masculine' in gender_text:
                    name_data['gender'] = 'm'
                elif 'feminine' in gender_text:
                    name_data['gender'] = 'f'
                else:
                    name_data['gender'] = gender_text # Fallback
        
        # Process Usage
        if usage:
            usage_span = name_span.find_next_sibling('span', class_='listusage')
            if usage_span:
                # Find all <a> tags within the usage span to get each usage part
                usage_links = usage_span.find_all('a')
                name_data['usage'] = [u.text.strip() for u in usage_links]
        
        # Process Description
        if desc:
            br_tag = parent_div.find('br')
            if br_tag:
                desc_parts = []
                current_node = br_tag.next_sibling
                while current_node:
                   
                    if isinstance(current_node, NavigableString):
                        desc_parts.append(str(current_node))
                    elif hasattr(current_node, 'get_text'):
                        desc_parts.append(current_node.get_text())
                    
                    current_node = current_node.next_sibling
                
                full_desc = ''.join(desc_parts).strip()
                if full_desc:
                    name_data['desc'] = full_desc
        
        # Add the collected data to the main dictionary
        results[name] = name_data
            
    return results


def scrape_names(pages: List[int], output_file_path: str, gender: bool = True, usage: bool = False, desc: bool = False, output_format: str = "csv"):
    """
    Scrapes name data from a list of pages and saves to a file in the specified format.
    
    Args:
        pages: A list of page numbers to scrape.
        output_file_path: The path to the output file (e.g., "names.csv").
        gender: Include gender information.
        usage: Include usage information.
        desc: Include description information.
        output_format: The format of the output file ("csv", "json", or "xml").
    """
    
    # This master dictionary will hold all data from all pages
    all_names_data: Dict[str, Dict[str, Any]] = {}

    print(f"Starting scrape for {len(pages)} pages...")
    
    for page_num in pages:
        print(f"Scraping page {page_num}...")
        html = get_names_page(page_num)
        if html:
            page_data = extract_names_from_page(html, gender, usage, desc)
            # Update the master dictionary
            all_names_data.update(page_data)
        else:
            print(f"Skipping page {page_num} (failed to fetch).")
    
    print(f"Scrape complete. Total unique names found: {len(all_names_data)}")
    print(f"Writing data to {output_file_path} as {output_format}...")

    # Write the aggregated data to the specified file format
    
    if output_format == "csv":
        # Define header based on boolean flags
        fieldnames = ['name']
        if gender: fieldnames.append('gender')
        if usage: fieldnames.append('usage')
        if desc: fieldnames.append('desc')
        
        with open(output_file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for name, data in all_names_data.items():
                row = {'name': name}
                if gender:
                    row['gender'] = data.get('gender')
                if usage:
                    # Join the list of usages into a single string for CSV
                    row['usage'] = ', '.join(data.get('usage', []))
                if desc:
                    row['desc'] = data.get('desc')
                writer.writerow(row)
                
    elif output_format == "json":
        with open(output_file_path, 'w', encoding='utf-8') as f:
            json.dump(all_names_data, f, indent=4, ensure_ascii=False)
            
    elif output_format == "xml":
        root = ET.Element('names')
        for name, data in all_names_data.items():
            name_elem = ET.SubElement(root, 'name', value=name)
            
            if gender and data.get('gender'):
                ET.SubElement(name_elem, 'gender').text = data.get('gender')
                
            if usage and data.get('usage'):
                # Create a separate <usage> tag for each item in the list
                for u in data.get('usage', []):
                    ET.SubElement(name_elem, 'usage').text = u
            
            if desc and data.get('desc'):
                ET.SubElement(name_elem, 'desc').text = data.get('desc')
                
        tree = ET.ElementTree(root)
        # Pretty-print the XML
        ET.indent(tree, space="  ")
        tree.write(output_file_path, encoding='utf-8', xml_declaration=True)
        
    else:
        print(f"Error: Unsupported output format '{output_format}'. Please use 'csv', 'json', or 'xml'.")

    print("File writing complete.")




def json_to_excel(json_file_path: str, excel_file_path: str):
    """
    Reads a JSON file produced by scrape_names and writes its contents
    to an Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "names"
    
    try:
        # Open and load the JSON file
        with open(json_file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        if not data:
            print(f"Warning: {json_file_path} is empty.")
            wb.save(excel_file_path)
            return
        
        headers = ['name']
        
        # Check the keys of the *first* item to see what columns to add
        # We assume all items in the file have the same structure
        first_item_keys = data[next(iter(data))].keys()
        
        if 'gender' in first_item_keys:
            headers.append('gender')
        if 'usage' in first_item_keys:
            headers.append('usage')
        if 'desc' in first_item_keys:
            headers.append('desc')
            
        # Write the dynamically created header row
        ws.append(headers)
            
        # Iterate over the JSON data
        for name, info in data.items():
            row = [name]
            if 'gender' in headers:
                row.append(info.get('gender'))
            if 'usage' in headers:
                # Join the list of usages into a single comma-separated string
                row.append(', '.join(info.get('usage', [])))
            if 'desc' in headers:
                row.append(info.get('desc'))
            
            # Append the row to the worksheet
            ws.append(row)
            
        # Save the workbook to the specified file
        wb.save(excel_file_path)
        print(f"Successfully converted {json_file_path} to {excel_file_path}")
        
    except FileNotFoundError:
        print(f"Error: The file {json_file_path} was not found.")
    except json.JSONDecodeError:
        print(f"Error: Could not decode JSON from {json_file_path}.")
    except Exception as e:
        print(f"An error occurred during JSON to Excel conversion: {e}")


def csv_to_excel(csv_file_path: str, excel_file_path: str):
    """
    Reads a CSV file and writes its contents to an Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "names"
    
    try:
        # Open the CSV file
        with open(csv_file_path, 'r', encoding='utf-8') as f:
            # Use csv.reader to read the file
            reader = csv.reader(f)
            # Iterate over each row in the CSV
            for row in reader:
                # Append the row directly to the worksheet
                ws.append(row)
                
        # Save the workbook
        wb.save(excel_file_path)
        print(f"Successfully converted {csv_file_path} to {excel_file_path}")

    except FileNotFoundError:
        print(f"Error: The file {csv_file_path} was not found.")
    except Exception as e:
        print(f"An error occurred during CSV to Excel conversion: {e}")


def xml_to_excel(xml_file_path: str, excel_file_path: str):
    """
    Reads an XML file produced by scrape_names and writes its contents
    to an Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "names"
    
    try:
        # Parse the XML file
        tree = ET.parse(xml_file_path)
        root = tree.getroot()
        
        headers = ['name']
        
        # Find the first <name> element to check its structure
        first_name_element = root.find('name')
        
        if first_name_element is not None:
            # Check for child tags in the *first* element
            if first_name_element.find('gender') is not None:
                headers.append('gender')
            if first_name_element.find('usage') is not None:
                headers.append('usage')
            if first_name_element.find('desc') is not None:
                headers.append('desc')
        else:
            print(f"Warning: {xml_file_path} contains no <name> elements.")
            wb.save(excel_file_path)
            return
            
        # Write the dynamically created header row
        ws.append(headers)
        # --- END FIX ---
        
        # Iterate over all <name> elements in the XML
        for name_elem in root.findall('name'):
            # --- FIX: DYNAMIC ROW CREATION ---
            row = [name_elem.get('value')] # Get 'name' from attribute
            
            if 'gender' in headers:
                gender_tag = name_elem.find('gender')
                row.append(gender_tag.text if gender_tag is not None else None)
            
            if 'usage' in headers:
                usage_tags = name_elem.findall('usage')
                row.append(', '.join([tag.text for tag in usage_tags if tag.text]))
            
            if 'desc' in headers:
                desc_tag = name_elem.find('desc')
                row.append(desc_tag.text if desc_tag is not None else None)
            
            # Append the dynamically created row
            ws.append(row)


        # Save the workbook
        wb.save(excel_file_path)
        print(f"Successfully converted {xml_file_path} to {excel_file_path}")
        
    except FileNotFoundError:
        print(f"Error: The file {xml_file_path} was not found.")
    except ET.ParseError:
        print(f"Error: Could not parse XML from {xml_file_path}.")
    except Exception as e:
        print(f"An error occurred during XML to Excel conversion: {e}")


if __name__ == "__main__":


    # --- Task 1: Scrape data and create files ---
    print("--- Running Task 1: Scraping Data ---")
    test_pages = [1, 2] # Scrape 2 pages for testing
    
    scrape_names(test_pages, "names_output.csv", gender=True, usage=True, desc=True, output_format="csv")
    scrape_names(test_pages, "names_output.json", gender=True, usage=True, desc=True, output_format="json")
    scrape_names(test_pages, "names_output.xml", gender=True, usage=True, desc=True, output_format="xml")
    
    
    # --- Task 2: Convert files to Excel ---
    print("\n--- Running Task 2: Converting to Excel ---")
    
    # Test JSON to Excel
    print("Testing JSON to Excel...")
    json_to_excel("names_output.json", "names_from_json.xlsx")
    
    # Test CSV to Excel
    print("Testing CSV to Excel...")
    csv_to_excel("names_output.csv", "names_from_csv.xlsx")

    # Test XML to Excel
    print("Testing XML to Excel...")
    xml_to_excel("names_output.xml", "names_from_xml.xlsx")
    
    print("\n--- All tests complete. Check output files. ---")
